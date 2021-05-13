#!/usr/bin/env python
# encoding: utf-8
"""
    @author: madgd
    @license: (C) Copyright 2020-2021 madgd. All Rights Reserved.
    @contact: madgdtju@gmail.com
    @software: 
    @file: excel_merger.py
    @time: 2020/10/26 3:53 下午
    @desc:
"""
import os
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../")
import argparse
import openpyxl
from utils.utils import copyLine, getCellValues, checkEmptyLine
from os.path import basename, dirname, abspath
import time


def excelMergeBySheet(excelFolder, outputPath="", headLines=1, sheetNum=1, sheetNameKey="",
                      allSheet=False, rmDup=False, nameKey=""):
    """
    input folder, merge all excels to 1
    :param excelFolder: input excel files‘ folder
    :param outputPath: default root folder of excelFolder
    :param headLines:
    :param sheetNum: if set, search sheet by num
    :param sheetNameKey: if set, search sheet by name key
    :param allSheet: if set true, merge all sheet by sheetNum
    :param rmDup: if set ture, remove dup line
    :param nameKey: if set, only merge excels that contains this key
    :return:
    """
    err = ""
    # name output by nameKey if it is set, else name it by first excel
    excelName = ""
    if nameKey:
        excelName = nameKey


    # read files
    try:
        files = os.listdir(excelFolder)
    except:
        err = "read folder err"
        return "", err
    absPath = os.path.abspath(excelFolder)
    headers = []
    mergedRowsList = []
    sheetNames = []
    # validation
    validationsBySheet = []
    # column_dims
    columnDimsBySheet = []
    # filter
    filterBySheet = []
    for file in files:
        # todo: search files recursively
        if not os.path.isdir("%s/%s" % (absPath, file)):
            if nameKey != "" and nameKey in file or nameKey == "":
                try:
                    print("%s/%s" % (absPath, file))
                    wb = openpyxl.load_workbook(filename="%s/%s" % (absPath, file))
                except: # not excel
                    print("%s not excel" % file)
                    continue
            else:
                continue
        else:
            continue

        if not excelName:
            excelName = ".".join(basename(file).split(".")[:-1]) + "_etc"

        # sheets to process
        tmpSheet = []
        if not allSheet:
            # get target sheet
            tmpSheetNames = wb.sheetnames
            # default
            targetSheet = wb[tmpSheetNames[0]]
            # if sheetNum set
            if sheetNum != 1:
                if type(sheetNum) is int and sheetNum > 1 and sheetNum <= len(tmpSheetNames):
                    targetSheet = wb[tmpSheetNames[sheetNum - 1]]
                else:
                    err = "sheetNum err"
                    return file, err
            # if sheetNameKey set
            if sheetNameKey:
                found = False
                for i in range(len(tmpSheetNames)):
                    if sheetNameKey in tmpSheetNames[i]:
                        found = True
                        targetSheet = wb[tmpSheetNames[i]]
                        break
                if not found:
                    err = "sheetNameKey not found"
                    return file, err
            tmpSheet.append(targetSheet)
        else:
            tmpSheet = wb.worksheets

        # process sheets
        for index in range(len(tmpSheet)):
            targetSheet = tmpSheet[index]
            # sheet name
            if index >= len(sheetNames):
                sheetName = targetSheet.title
                sheetNames.append(sheetName)
            # sheet header
            if index >= len(headers):
                tmpHeader = []
                for i in range(headLines):
                    tmpHeader.append(targetSheet[i+1])
                headers.append(tmpHeader)
                validationsBySheet.append(targetSheet.data_validations.dataValidation)
                columnDimsBySheet.append(targetSheet.column_dimensions)
                filterBySheet.append(targetSheet.auto_filter)
            # rows
            if index >= len(mergedRowsList):
                mergedRowsList.append([])
            allRows = targetSheet.rows
            for i in range(headLines):
                next(allRows)
            for row in allRows:
                mergedRowsList[index].append(row)

    # rm dup line
    if rmDup:
        for i in range(len(mergedRowsList)):
            tmp = []
            s = set()
            for row in mergedRowsList[i]:
                key = "".join([str(i) for i in getCellValues(row)])
                if key in s:
                    continue
                s.add(key)
                tmp.append(row)
            mergedRowsList[i] = tmp


    # output
    if excelFolder[-1] == "/":
        excelFolder = excelFolder[:-1]
    rootFolder = abspath(dirname(excelFolder))
    if outputPath == "":
        outputPath = rootFolder + "/%s_merged_%s.xlsx" % (excelName, time.strftime("%Y_%m_%d-%H_%M", time.localtime()))
    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    for index in range(len(sheetNames)):
        sheet = workbook.create_sheet(sheetNames[index])
        # validation
        sheet.data_validations.dataValidation = validationsBySheet[index]
        # col_dims
        sheet.column_dimensions = columnDimsBySheet[index]
        # filter
        sheet.auto_filter = filterBySheet[index]
        curr = 0
        for line in headers[index]:
            copyLine(sheet, line, curr)
            curr += 1
        for line in mergedRowsList[index]:
            # remove empty line
            if not checkEmptyLine(line):
                copyLine(sheet, line, curr)
                curr += 1
    workbook.save(outputPath)

    return outputPath, err

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='merge excel.')
    parser.add_argument('excelFolder', metavar='input', type=str,
                        help='input excel folder')
    parser.add_argument('-o', metavar='output', type=str, default="",
                        help='output path, default input root folder')
    parser.add_argument('-l', metavar='headLines', type=int, default=1,
                        help='header lines, default 1')
    parser.add_argument('-s', metavar='sheetNum', type=int, default=1,
                        help='which sheet to merge, default 1')
    parser.add_argument('-k', metavar='sheetNameKey', type=str, default="",
                        help='which sheet to merge, search by key. won`t work if not set')
    parser.add_argument('-a', action='store_true',
                        help='if set, merge all sheets by num. default not')
    parser.add_argument('-d', action='store_true',
                        help='if set, remove dup line. default not')
    parser.add_argument('-n', metavar='excelNameKey', type=str, default="",
                        help='if set, only merge excels that contains this key')
    args = parser.parse_args()

    outputPath, err = excelMergeBySheet(args.excelFolder, args.o, headLines=args.l, sheetNum=args.s, sheetNameKey=args.k, \
                      allSheet=args.a, rmDup=args.d, nameKey=args.n)
    if err:
        print("merge failed because: %s(%s)" % (err, outputPath))
    else:
        print("merge finished! output at %s" % outputPath)