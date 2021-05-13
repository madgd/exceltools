#!/usr/bin/env python3
# encoding: utf-8
"""
    @author: madgd
    @license: (C) Copyright 2020-2021 madgd. All Rights Reserved.
    @contact: madgdtju@gmail.com
    @software:
    @file: excel_spliter.py
    @time: 2020/9/30 5:05 下午
    @desc: split excel by column value
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../")
import openpyxl
from utils.utils import titleToNumber, filterByList, getCellValues, copyLine, findColNumByName
from os.path import basename, dirname
import argparse
import time

sep = "$\001$"
def excelSplitBySheet(excelPath, outputPath='', columnLabels="A", headLines=1, sheetNum=1, sheetNameKey="", allSheet=False, *sheetLabels):
    """
    输入excel文件路径，按参数拆分excel表，并返回多个文件对象
    :param excelPath: 输入路径
    :param outputPath: 输出路径
    :param columnLabels: 拆分依据列。默认拆分第一个sheet，拆分依据可为多列, 用英文","隔开, 有顺序
    :param headLines: 跳过表头行数
    :param sheetNum: 第几个sheet
    :param sheetNameKey: 按sheet名称的关键词找拆分表，默认为空。
    :param allSheet: 是否根据第一个sheet的拆分依据，自动找到其他表中相应的列，也进行拆分
    :param sheetLabels: 按传入的labels，同时对不同的sheet进行拆分。生成m*n个表。目前不生效
    :return: outputPath
    """
    err = ""
    # read excel
    wb = openpyxl.load_workbook(filename=excelPath)
    excelName = ".".join(basename(excelPath).split(".")[:-1])
    # print(wb.sheet_names())


    # get target sheet
    sheetNames = wb.sheetnames
    # default
    targetSheet = wb[sheetNames[0]]
    # if sheetNum set
    if sheetNum != 1:
        if type(sheetNum) is int and sheetNum > 1 and sheetNum <= len(sheetNames):
            targetSheet = wb[sheetNames[sheetNum - 1]]
        else:
            err = "sheetNum err"
            return outputPath, err
    # if sheetNameKey set
    if sheetNameKey:
        found = False
        for i in range(len(sheetNames)):
            if sheetNameKey in sheetNames[i]:
                found = True
                targetSheet = wb[sheetNames[i]]
                break
        if not found:
            err = "sheetNameKey not found"
            return outputPath, err
    # print(targetSheet.name, targetSheet.nrows, targetSheet.ncols)


    # sheet header
    header = []
    for i in range(headLines):
        header.append(targetSheet[i+1])
        # header.append(getCellValues(targetSheet.row(i)))

    # split by cols
    # find colnames
    cols2split = columnLabels.split(",")
    cols2splitNum = [titleToNumber(i) - 1 for i in cols2split]
    targetColNames = filterByList(header[-1], cols2splitNum)
    # print(targetColNames)
    # print(findColNumByName(header[-1], targetColnames))
    sheets = [targetSheet]
    # if split all sheet
    if allSheet:
        sheets = wb.worksheets
    else:
        sheetNames = [targetSheet.title]
    rowGroupsBySheet = {}
    headerBySheet = {}
    # validations
    validationsBySheet = {}
    # column_dims
    columnDimsBySheet = {}
    # filter
    filterBySheet = {}
    for sheet in sheets:
        sheetName = sheet.title
        # validations
        validationsBySheet[sheetName] = sheet.data_validations.dataValidation
        # column_dims
        columnDimsBySheet[sheetName] = sheet.column_dimensions
        # filter
        filterBySheet[sheetName] = sheet.auto_filter
        # find cols2splitNum by targetColnames
        tmpHeader = []
        for i in range(headLines):
            tmpHeader.append((sheet[i+1]))
        # print(tmpHeader)
        headerBySheet[sheetName] = tmpHeader
        tmpCols2splitNum = findColNumByName(tmpHeader[-1], targetColNames)
        # print(tmpCols2splitNum)

        allRows = sheet.rows
        for i in range(headLines):
            next(allRows)
        for row in allRows:
            dicKey = sep.join([str(i) for i in getCellValues(filterByList(row, tmpCols2splitNum))])
            if dicKey in rowGroupsBySheet:
                if sheetName in rowGroupsBySheet[dicKey]:
                    rowGroupsBySheet[dicKey][sheetName].append(row)
                else:
                    rowGroupsBySheet[dicKey][sheetName] = [row]
            else:
                rowGroupsBySheet[dicKey] = {sheetName: [row]}



    # save output excels
    if outputPath == "":
        outputPath = dirname(excelPath) + "/%s_split_%s" % (excelName, time.strftime("%Y_%m_%d-%H_%M", time.localtime()))
    if not os.path.exists(outputPath):
        os.mkdir(outputPath)
    for k in rowGroupsBySheet:
        workbook = openpyxl.Workbook()
        del workbook['Sheet']
        for sheetName in sheetNames:
            sheet = workbook.create_sheet(sheetName)
            if sheetName not in rowGroupsBySheet[k]:
                rowGroupsBySheet[k][sheetName] = []
            curr = 0
            # header
            for line in headerBySheet[sheetName]:
                copyLine(sheet, line, curr)
                curr += 1
            for line in rowGroupsBySheet[k][sheetName]:
                copyLine(sheet, line, curr)
                curr += 1
            # validation
            sheet.data_validations.dataValidation = validationsBySheet[sheetName]
            # col_dims
            sheet.column_dimensions = columnDimsBySheet[sheetName]
            # filter
            sheet.auto_filter = filterBySheet[sheetName]
        # return
        workbook.save("%s/%s_%s.xlsx" % (outputPath, excelName, k))

    return outputPath, err

def main():
    print("in main")
    # excelSplitBySheet()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='split excel.')
    parser.add_argument('excelPath', metavar='input', type=str,
                        help='input excel')
    parser.add_argument('-o', metavar='output', type=str, default="",
                        help='output path, default input dic')
    parser.add_argument('-c', metavar='columnLabels', type=str, default="A",
                        help='column labels in alphabat to split, multi label sep by ",". default "A"')
    parser.add_argument('-l', metavar='headLines', type=int, default=1,
                        help='header lines, default 1')
    parser.add_argument('-s', metavar='sheetNum', type=int, default=1,
                        help='which sheet to split, default 1')
    parser.add_argument('-k', metavar='sheetNameKey', type=str, default="",
                        help='which sheet to split, search by key. won`t work if not set')
    parser.add_argument('-a', action='store_true',
                        help='if set, split all sheets. default not')
    args = parser.parse_args()

    excelSplitBySheet(args.excelPath, args.o, columnLabels=args.c, headLines=args.l, sheetNum=args.s,\
                      sheetNameKey=args.k ,allSheet=args.a)
    print("split finished!")